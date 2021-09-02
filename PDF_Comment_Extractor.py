import fitz
import os
import openpyxl
folder_path ="D:\pdf"
all_files = []
for dirpath, dirnames, filenames in os.walk(folder_path):
    for filename in [f for f in filenames if f.endswith(".pdf")]:
        all_files.append(os.path.join(dirpath, filename))

def extractComment():
     r = 1
     for fileName in all_files:
         doc = fitz.open(fileName)
         sh.cell(row=r + 1, column=1).value = fileName
         for i in range(doc.pageCount):
                 page = doc[i]
                 for annot in page.annots():
                     comment = annot.info["content"]
                     author = annot.info["title"]
                     sh.cell(row=r + 1, column=4).value = comment
                     sh.cell(row=r + 1, column=3).value = author
                     sh.cell(row=r + 1, column=2).value = i+1
                    # sh.cell(row=r + 1, column=1).value = fileName
                     r=r+1

if __name__ == '__main__':
    wk = openpyxl.Workbook()
    sh = wk.active
    sh.title = "CommentRecord"
    sh['A1'].value = "Location & File Name"
    sh['B1'].value = "Page No."
    sh['C1'].value = "Author Name"
    sh['D1'].value = "Comments"
    extractComment()
    wk.save("D:\pdf\CommentRecord.xlsx")
